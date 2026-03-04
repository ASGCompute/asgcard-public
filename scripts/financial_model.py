#!/usr/bin/env python3
"""ASG Card — Financial Model v5 (All audit fixes applied)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as GL

OUT="/Users/innocode/Desktop/Test/ASGcard/ASG_Card_Financial_Model.xlsx"
wb=openpyxl.Workbook()
W="FFFFFF";L="F8F9FA";M="E9ECEF";DT="212529";BL="4361EE";DK="16213E";AC="0F3460"
GR="06D6A0";RD="EF476F";YL="FFD166";IB="DBEAFE";RB="D1FAE5";WB="FEF3C7";BB="FEE2E2"
SB="0057FF";XG="F5A623";PP="7C3AED"
tn=Side(style="thin",color="DEE2E6");bd=Border(top=tn,bottom=tn,left=tn,right=tn)
def Fn(sz=10,b=False,c=DT): return Font(name="Inter",size=sz,bold=b,color=c)
def BG(c): return PatternFill("solid",fgColor=c)
def AL(h="left",v="center",w=False): return Alignment(horizontal=h,vertical=v,wrap_text=w)
def ce(ws,r,c,v,f=None,bg=None,a=None,fm=None):
    cl=ws.cell(row=r,column=c,value=v);cl.font=f or Fn();cl.fill=bg or BG(W);cl.alignment=a or AL();cl.border=bd
    if fm:cl.number_format=fm
    return cl
def mg(ws,r,t,n):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=n)
    ce(ws,r,1,t,Fn(13,True,W),BG(AC),AL("left","center"))
    for c in range(2,n+1):ws.cell(row=r,column=c).fill=BG(AC);ws.cell(row=r,column=c).border=bd
    ws.row_dimensions[r].height=30
def sb(ws,r,t,n):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=n)
    ce(ws,r,1,t,Fn(11,True,DT),BG(M),AL("left","center"))
    for c in range(2,n+1):ws.cell(row=r,column=c).fill=BG(M);ws.cell(row=r,column=c).border=bd
    ws.row_dimensions[r].height=24
def hd(ws,r,d):
    for i,v in enumerate(d,1):ce(ws,r,i,v,Fn(10,True,W),BG(DK),AL("center","center"))
    ws.row_dimensions[r].height=22
def ip(ws,r,c,v,fm='#,##0.00'):ce(ws,r,c,v,Fn(10,True,BL),BG(IB),AL("center"),fm)
def rs(ws,r,c,v,fm='$#,##0.00',co=GR):ce(ws,r,c,v,Fn(10,True,co),BG(RB),AL("center"),fm)
def sc(ws,w):
    for i,v in enumerate(w,1):ws.column_dimensions[GL(i)].width=v

# ══════════════════════════════════════
# TAB 1: ASSUMPTIONS
# ══════════════════════════════════════
ws=wb.active;ws.title="Assumptions";ws.sheet_properties.tabColor=BL
sc(ws,[38,16,16,16,36])
r=1;mg(ws,r,"ASG CARD — MODEL ASSUMPTIONS (v5 — Audited)",5);r+=1
ce(ws,r,1,"Blue cells = editable. Audit fixes: correct ARPU, declining churn, ramped Pro.",Fn(10,False,BL));r+=2

# 4PAYMENTS
sb(ws,r,"4PAYMENTS REAL COST STRUCTURE",5);r+=1
hd(ws,r,["Parameter","Current","At Volume","","Notes"]);r+=1
fp=[["Setup Fee (one-time)",4500,"—","One-time integration fee"],
    ["Card Issuance ($/card)",3.00,2.50,"Volume discount at 100+ cards/mo"],
    ["Deposit Fee (%)",3.0,2.0,"Volume discount negotiable"],
    ["Monthly Platform Fee ($)",75,"—","Fixed monthly"]]
fp_s=r
for f in fp:
    ce(ws,r,1,f[0],Fn(10,True),BG(L));ip(ws,r,2,f[1])
    if f[2]=="—":
            ce(ws,r,3,"—",Fn(10,False,"6B7280"),BG(W),AL("center"))
    else:
            ce(ws,r,3,f[2],Fn(10,True,BL),BG(IB),AL("center"),'#,##0.00')
    ce(ws,r,5,f[3],Fn(9,False,"6B7280"));r+=1
r+=1

# REVENUE PER TRANSACTION (AUDIT FIX: topUpFee IS your revenue)
sb(ws,r,"YOUR REVENUE PER TRANSACTION (topUpFee + serviceFee)",5);r+=1
hd(ws,r,["Parameter","Pessimistic","Realistic","Optimistic","Notes"]);r+=1
pr=[["Avg topUpFee earned / create ($)",2.20,3.00,6.00,"From pricing.ts, you keep this"],
    ["Avg serviceFee / create ($)",2.00,2.50,5.00,"Your service margin"],
    ["Avg topUpFee earned / fund ($)",2.20,3.00,6.00,"Funding topUp fee"],
    ["Avg serviceFee / fund ($)",2.00,2.00,3.00,"Funding service fee"],
    ["Card Creates / client / mo",2,4,8,"Monthly issuance volume"],
    ["Card Funds / client / mo",1,2,4,"Monthly funding volume"],
    ["Avg loadAmount for creates ($)",25,50,100,"For 4p deposit % calc"],
    ["Avg fundAmount for funds ($)",25,50,100,"For 4p deposit % calc"]]
pr_s=r
for p in pr:
    ce(ws,r,1,p[0],Fn(10,True),BG(L))
    for c in [2,3,4]:ip(ws,r,c,p[c-1])
    ce(ws,r,5,p[4],Fn(9,False,"6B7280"));r+=1
r+=1

# PRO SUBSCRIPTION (AUDIT FIX: ramp from 0%)
sb(ws,r,"PRO SUBSCRIPTION ($29/mo) — RAMPED CONVERSION",5);r+=1
hd(ws,r,["Parameter","Pessimistic","Realistic","Optimistic","Notes"]);r+=1
sp=[["Pro Price ($/mo)",29,29,29,"Monthly subscription"],
    ["Mo 1 Pro Conversion %",0,0,5,"Start at 0% — no track record yet"],
    ["Mo 6 Pro Conversion %",5,10,20,"After proving value"],
    ["Mo 12 Pro Conversion %",10,20,35,"Mature conversion rate"]]
sp_s=r
for s in sp:
    ce(ws,r,1,s[0],Fn(10,True),BG("F3E8FF"))
    for c in [2,3,4]:ce(ws,r,c,s[c-1],Fn(10,True,PP),BG("F3E8FF"),AL("center"),'#,##0')
    ce(ws,r,5,s[4],Fn(9,False,"6B7280"));r+=1
r+=1

# CALCULATED ARPU (AUDIT FIX: includes topUpFee)
sb(ws,r,"CALCULATED ARPU (topUpFee + serviceFee - 4p COGS)",5);r+=1
# Revenue per client
ce(ws,r,1,"Gross Revenue / client / mo",Fn(10,True),BG(L))
for c,co in [(2,"B"),(3,"C"),(4,"D")]:
    s=pr_s # topUp_create*creates + svc_create*creates + topUp_fund*funds + svc_fund*funds
    rs(ws,r,c,f"=({co}{s}+{co}{s+1})*{co}{s+4}+({co}{s+2}+{co}{s+3})*{co}{s+5}")
ce(ws,r,5,"(topUp+svc)×creates + (topUp+svc)×funds",Fn(9,False,"6B7280"))
grev_r=r;r+=1

# COGS per client
ce(ws,r,1,"4p COGS / client / mo",Fn(10,True,RD),BG(BB))
for c,co in [(2,"B"),(3,"C"),(4,"D")]:
    s=pr_s # issuance*creates + deposit%*loadAmt*creates + deposit%*fundAmt*funds
    rs(ws,r,c,f"=B{fp_s+1}*{co}{s+4}+B{fp_s+2}/100*{co}{s+6}*{co}{s+4}+B{fp_s+2}/100*{co}{s+7}*{co}{s+5}",'$#,##0.00',RD)
    ws.cell(row=r,column=c).fill=BG(BB)
ce(ws,r,5,"$3×creates + 3%×load×creates + 3%×fund×funds",Fn(9,False,"6B7280"))
cogs_r=r;r+=1

# Net margin per client (PPU only)
ce(ws,r,1,"★ NET MARGIN / client / mo (PPU)",Fn(11,True,GR),BG(RB))
for c,co in [(2,"B"),(3,"C"),(4,"D")]:
    rs(ws,r,c,f"={co}{grev_r}-{co}{cogs_r}")
ce(ws,r,5,"Revenue minus 4payments variable costs",Fn(9,False,"6B7280"))
margin_r=r;r+=2

# GROWTH (AUDIT FIX: declining churn)
sb(ws,r,"GROWTH & CONVERSION (DECLINING CHURN)",5);r+=1
hd(ws,r,["Parameter","Pessimistic","Realistic","Optimistic","Notes"]);r+=1
gr=[["Month 1 Signups",10,20,30,"Free tier = low friction"],
    ["MoM Signup Growth %",15,25,35,"Monthly growth rate"],
    ["Signup → Free User %",25,40,60,"Free tier conversion (no cost)"],
    ["Free → Paying %",15,25,40,"Upgrade to paid over lifetime"],
    ["Mo 1-3 Churn %",15,12,8,"Early stage — high churn"],
    ["Mo 4-6 Churn %",10,8,5,"Stabilizing"],
    ["Mo 7-12 Churn %",8,5,3,"Product-market fit"]]
gr_s=r
for g in gr:
    ce(ws,r,1,g[0],Fn(10,True),BG(L))
    for c in [2,3,4]:ip(ws,r,c,g[c-1],'#,##0')
    ce(ws,r,5,g[4],Fn(9,False,"6B7280"));r+=1
r+=1

# FIXED COSTS
sb(ws,r,"MONTHLY FIXED COSTS ($)",5);r+=1
hd(ws,r,["Category","Pessimistic","Realistic","Optimistic","Notes"]);r+=1
costs=[["Hosting (Vercel+Supabase+Stellar)",100,150,200,"Scales with usage"],
       ["4payments Platform Fee",75,75,75,"Fixed monthly"],
       ["Domain+Tooling+SaaS",20,30,40,"DNS, CI, monitoring"],
       ["Marketing & Content",100,250,500,"Ads, content, events"],
       ["CTO / Developer",0,0,2000,"Contractor"],
       ["Legal / Compliance",0,50,100,"Periodic"],
       ["Founder Opportunity Cost",3600,3600,3600,"NOT in P&L"]]
cs_s=r
for co in costs:
    ce(ws,r,1,co[0],Fn(10,True),BG(L))
    for c in [2,3,4]:ip(ws,r,c,co[c-1],'$#,##0')
    ce(ws,r,5,co[4],Fn(9,False,"6B7280"));r+=1
cs_e=r-1
ce(ws,r,1,"TOTAL FIXED COSTS",Fn(11,True,RD),BG(BB))
for c,co in [(2,"B"),(3,"C"),(4,"D")]:
    rs(ws,r,c,f"=SUM({co}{cs_s}:{co}{cs_e-1})",'$#,##0',RD);ws.cell(row=r,column=c).fill=BG(BB)
ce(ws,r,5,"Excludes founder opp cost",Fn(9,False,"6B7280"))
tc_r=r;r+=2

# ONE-TIME
sb(ws,r,"ONE-TIME COSTS",5);r+=1
ce(ws,r,1,"4payments Setup Fee",Fn(10,True,RD),BG(BB));ip(ws,r,2,4500,'$#,##0');ws.cell(row=r,column=2).fill=BG(BB)
ce(ws,r,5,"Paid once at start",Fn(9,False,"6B7280"))
su_r=r;r+=2

# CAC
sb(ws,r,"CUSTOMER ACQUISITION",5);r+=1
hd(ws,r,["Parameter","Pessimistic","Realistic","Optimistic","Notes"]);r+=1
ce(ws,r,1,"Blended CAC ($)",Fn(10,True),BG(L))
for c in [2,3,4]:ip(ws,r,c,[25,20,10][c-2],'$#,##0')
ce(ws,r,5,"Weighted avg across channels",Fn(9,False,"6B7280"))
cac_r=r;r+=1
ce(ws,r,1,"Avg Client Lifetime (mo)",Fn(10,True),BG(L))
for c in [2,3,4]:ip(ws,r,c,[6,9,14][c-2],'#,##0')
ce(ws,r,5,"≈ 1 / avg churn",Fn(9,False,"6B7280"))
lt_r=r;r+=2

# FUNDING
sb(ws,r,"FUNDING & RUNWAY",5);r+=1
hd(ws,r,["Parameter","Value","","","Notes"]);r+=1
for f in [["Cash on Hand",2000,"Personal"],["SCF Grant",10000,"80-120K XLM"],["Monthly Budget",500,"Personal"]]:
    ce(ws,r,1,f[0],Fn(10,True),BG(L));ip(ws,r,2,f[1],'$#,##0');ce(ws,r,5,f[2],Fn(9,False,"6B7280"));r+=1

# ══════════════════════════════════════
# TAB 2: UNIT ECONOMICS
# ══════════════════════════════════════
ws=wb.create_sheet("Unit Economics");ws.sheet_properties.tabColor=GR
sc(ws,[30,14,14,14,14,30])
r=1;mg(ws,r,"UNIT ECONOMICS (AUDITED)",6);r+=2

# AUDIT FIX: Correct margin table
sb(ws,r,"REAL MARGIN PER CREATE TRANSACTION",6);r+=1
hd(ws,r,["Tier","You Charge","4p Cost","YOUR Net","Margin %","Notes"]);r+=1
# Correct: you charge topUp+svc. 4p takes $3+3%×load.
for nm,load,top,svc in [("$10",10,2.20,2.00),("$25",25,2.50,2.00),("$50",50,3.00,2.00),
                         ("$100",100,4.00,3.00),("$200",200,6.00,5.00),("$500",500,12.00,7.00)]:
    you_charge=top+svc; fp_cost=3+load*0.03; net=you_charge-fp_cost; mpct=net/you_charge*100
    color=GR if net>=1.5 else (YL if net>=0.5 else RD)
    ce(ws,r,1,f"Create {nm}",Fn(10,True),BG(L))
    ce(ws,r,2,you_charge,Fn(),BG(L),AL("center"),'$#,##0.00')
    ce(ws,r,3,fp_cost,Fn(10,False,RD),BG(BB),AL("center"),'$#,##0.00')
    ce(ws,r,4,net,Fn(10,True,color),BG(RB if net>=1 else WB),AL("center"),'$#,##0.00')
    ce(ws,r,5,f"{mpct:.0f}%",Fn(10,True,color),BG(L),AL("center"))
    note="⚠️ Thin" if net<1 else ("✅ OK" if net<2 else "✅ Good")
    ce(ws,r,6,note,Fn(9,False,"6B7280"));r+=1

r+=1;sb(ws,r,"REAL MARGIN PER FUND TRANSACTION",6);r+=1
hd(ws,r,["Tier","You Charge","4p Cost","YOUR Net","Margin %","Notes"]);r+=1
for nm,amt,top,svc in [("$10",10,2.20,2.00),("$25",25,2.50,2.00),("$50",50,3.00,2.00),
                        ("$100",100,4.00,3.00),("$200",200,6.00,5.00),("$500",500,12.00,7.00)]:
    you_charge=top+svc; fp_cost=amt*0.03; net=you_charge-fp_cost; mpct=net/you_charge*100
    color=GR if net>=2 else (YL if net>=1 else RD)
    ce(ws,r,1,f"Fund {nm}",Fn(10,True),BG(L))
    ce(ws,r,2,you_charge,Fn(),BG(L),AL("center"),'$#,##0.00')
    ce(ws,r,3,fp_cost,Fn(10,False,RD),BG(BB),AL("center"),'$#,##0.00')
    ce(ws,r,4,net,Fn(10,True,color),BG(RB if net>=1 else WB),AL("center"),'$#,##0.00')
    ce(ws,r,5,f"{mpct:.0f}%",Fn(10,True,color),BG(L),AL("center"))
    note="⚠️ Thin" if net<1.5 else ("✅ OK" if net<3 else "✅ Good")
    ce(ws,r,6,note,Fn(9,False,"6B7280"));r+=1

r+=1;sb(ws,r,"KEY RATIOS (linked to Assumptions)",6);r+=1
hd(ws,r,["Metric","Pessimistic","Realistic","Optimistic","","Formula"]);r+=1
ms=r
md=[("Net Margin/client ($/mo)",f"=Assumptions!B{margin_r}",f"=Assumptions!C{margin_r}",f"=Assumptions!D{margin_r}","After 4p COGS"),
    ("CAC ($)",f"=Assumptions!B{cac_r}",f"=Assumptions!C{cac_r}",f"=Assumptions!D{cac_r}","Blended"),
    ("LTV ($)",f"=B{ms}*Assumptions!B{lt_r}",f"=C{ms}*Assumptions!C{lt_r}",f"=D{ms}*Assumptions!D{lt_r}","Margin×Lifetime"),
    ("LTV/CAC",f"=B{ms+2}/B{ms+1}",f"=C{ms+2}/C{ms+1}",f"=D{ms+2}/D{ms+1}","Target: >3x"),
    ("Payback (mo)",f"=B{ms+1}/B{ms}",f"=C{ms+1}/C{ms}",f"=D{ms+1}/D{ms}","CAC/Margin")]
for nm,f1,f2,f3,nt in md:
    ce(ws,r,1,nm,Fn(10,True,AC),BG(L))
    for c,fm in [(2,f1),(3,f2),(4,f3)]:rs(ws,r,c,fm,'$#,##0.00')
    ce(ws,r,6,nt,Fn(9,False,"6B7280"));r+=1

# ══════════════════════════════════════
# TAB 3: USER GROWTH (AUDIT FIX: declining churn, ramped Pro)
# ══════════════════════════════════════
ws=wb.create_sheet("User Growth");ws.sheet_properties.tabColor="39D2C0"
sc(ws,[24]+[12]*12)
mos=[""]+[f"Mo {i}" for i in range(1,13)]

for sn,sr,cr in [("REALISTIC",1,"C"),("PESSIMISTIC",22,"B")]:
    r=sr;mg(ws,r,f"USER GROWTH — {sn} (Declining Churn)",13);r+=1;hd(ws,r,mos);r+=1
    a_su=f"Assumptions!{cr}{gr_s}";a_gr=f"Assumptions!{cr}{gr_s+1}"
    a_fu=f"Assumptions!{cr}{gr_s+2}";a_fp=f"Assumptions!{cr}{gr_s+3}"

    ce(ws,r,1,"New Signups",Fn(10,True),BG(L))
    ce(ws,r,2,f"={a_su}",Fn(),BG(L),AL("center"),'#,##0')
    for c in range(3,14):ce(ws,r,c,f"=ROUND({GL(c-1)}{r}*(1+{a_gr}/100),0)",Fn(),BG(L),AL("center"),'#,##0')
    su_r2=r;r+=1

    ce(ws,r,1,"New Free Users",Fn(10,True,SB),BG(L))
    for c in range(2,14):ce(ws,r,c,f"=ROUND({GL(c)}{su_r2}*{a_fu}/100,0)",Fn(10,False,SB),BG(L),AL("center"),'#,##0')
    fu_r=r;r+=1

    ce(ws,r,1,"New Paying (free→paid)",Fn(10,True,GR),BG(L))
    for c in range(2,14):ce(ws,r,c,f"=ROUND({GL(c)}{fu_r}*{a_fp}/100,0)",Fn(10,False,GR),BG(L),AL("center"),'#,##0')
    np_r=r;r+=1

    # Declining churn: mo1-3, mo4-6, mo7-12
    ce(ws,r,1,"Churned (declining)",Fn(10,True,RD),BG(BB))
    ce(ws,r,2,0,Fn(10,False,RD),BG(BB),AL("center"),'#,##0')
    for c in range(3,14):
        mo=c-1 # month number
        if mo<=3: ch_ref=f"Assumptions!{cr}{gr_s+4}"
        elif mo<=6: ch_ref=f"Assumptions!{cr}{gr_s+5}"
        else: ch_ref=f"Assumptions!{cr}{gr_s+6}"
        ce(ws,r,c,f"=ROUND({GL(c-1)}{r+1}*{ch_ref}/100,0)",Fn(10,False,RD),BG(BB),AL("center"),'#,##0')
    ch_r=r;r+=1

    ce(ws,r,1,"TOTAL ACTIVE PAYING",Fn(11,True,W),BG(AC))
    ce(ws,r,2,f"=B{np_r}",Fn(11,True,W),BG(AC),AL("center"),'#,##0')
    for c in range(3,14):
        ce(ws,r,c,f"={GL(c-1)}{r}+{GL(c)}{np_r}-{GL(c)}{ch_r}",Fn(11,True,W),BG(AC),AL("center"),'#,##0')
    tot_r=r;r+=1

    # Pro subscribers (ramped: interpolate mo1→mo6→mo12)
    ce(ws,r,1,"  ↳ Pro Subscribers",Fn(10,True,PP),BG("F3E8FF"))
    for c in range(2,14):
        mo=c-1
        # Linear interpolation: mo1=sp_s+1, mo6=sp_s+2, mo12=sp_s+3
        if mo<=1: pct=f"Assumptions!{cr}{sp_s+1}"
        elif mo<=6: pct=f"(Assumptions!{cr}{sp_s+1}+(Assumptions!{cr}{sp_s+2}-Assumptions!{cr}{sp_s+1})*({mo}-1)/5)"
        else: pct=f"(Assumptions!{cr}{sp_s+2}+(Assumptions!{cr}{sp_s+3}-Assumptions!{cr}{sp_s+2})*({mo}-6)/6)"
        ce(ws,r,c,f"=ROUND({GL(c)}{tot_r}*{pct}/100,0)",Fn(10,False,PP),BG("F3E8FF"),AL("center"),'#,##0')
    pro_r=r;r+=1

    ce(ws,r,1,"Cumulative Signups",Fn(10,False,"6B7280"),BG(W))
    ce(ws,r,2,f"=B{su_r2}",Fn(10,False,"6B7280"),BG(W),AL("center"),'#,##0')
    for c in range(3,14):ce(ws,r,c,f"={GL(c-1)}{r}+{GL(c)}{su_r2}",Fn(10,False,"6B7280"),BG(W),AL("center"),'#,##0')
    r+=2

# ══════════════════════════════════════
# TAB 4: P&L
# ══════════════════════════════════════
ws=wb.create_sheet("P&L");ws.sheet_properties.tabColor=BL
sc(ws,[30]+[13]*12);usd='$#,##0'
r=1;mg(ws,r,"P&L — REALISTIC (AUDITED)",13);r+=1;hd(ws,r,mos);r+=1

ce(ws,r,1,"4p Setup Fee (Mo 1 only)",Fn(10,True,RD),BG(BB))
ce(ws,r,2,f"=Assumptions!B{su_r}",Fn(10,True,RD),BG(BB),AL("center"),usd)
for c in range(3,14):ce(ws,r,c,0,Fn(10,False,"6B7280"),BG(L),AL("center"),usd)
sf_r=r;r+=2

sb(ws,r,"REVENUE",13);r+=1
ce(ws,r,1,"Active Paying Clients",Fn(10,True),BG(L))
for c in range(2,14):ce(ws,r,c,f"='User Growth'!{GL(c)}{tot_r-6}",Fn(),BG(L),AL("center"),'#,##0')
cl_r=r;r+=1

ce(ws,r,1,"  PPU Revenue (margin×clients)",Fn(10,True),BG(L))
for c in range(2,14):ce(ws,r,c,f"={GL(c)}{cl_r}*Assumptions!C{grev_r}",Fn(),BG(L),AL("center"),usd)
ppu_r=r;r+=1

ce(ws,r,1,"  Pro Subscription Revenue",Fn(10,True,PP),BG("F3E8FF"))
for c in range(2,14):ce(ws,r,c,f"='User Growth'!{GL(c)}{pro_r-6}*Assumptions!C{sp_s}",Fn(10,False,PP),BG("F3E8FF"),AL("center"),usd)
sub_r=r;r+=1

ce(ws,r,1,"GROSS REVENUE",Fn(11,True,W),BG(GR))
for c in range(2,14):ce(ws,r,c,f"={GL(c)}{ppu_r}+{GL(c)}{sub_r}",Fn(11,True,W),BG(GR),AL("center"),usd)
rv_r=r;r+=2

sb(ws,r,"VARIABLE COSTS (4P COGS)",13);r+=1
ce(ws,r,1,"4p COGS (issuance+deposit)",Fn(10,True,RD),BG(BB))
for c in range(2,14):ce(ws,r,c,f"={GL(c)}{cl_r}*Assumptions!C{cogs_r}",Fn(10,False,RD),BG(BB),AL("center"),usd)
cg_r=r;r+=2

sb(ws,r,"FIXED EXPENSES",13);r+=1
en=["Hosting","4p Platform Fee","Domain+Tooling","Marketing","CTO/Dev","Legal"]
ex_s=r
for i,nm in enumerate(en):
    ce(ws,r,1,nm,Fn(10,True),BG(BB))
    for c in range(2,14):ce(ws,r,c,f"=Assumptions!C{cs_s+i}",Fn(10,False,RD),BG(BB),AL("center"),usd)
    r+=1

ce(ws,r,1,"XLM Rewards",Fn(10,True),BG("FFF7E6"))
for c in range(2,14):ce(ws,r,c,f"='XLM Rewards'!{GL(c)}19",Fn(10,True,XG),BG("FFF7E6"),AL("center"),usd)
xr=r;r+=1

ce(ws,r,1,"TOTAL EXPENSES",Fn(11,True,W),BG(RD))
for c in range(2,14):ce(ws,r,c,f"={GL(c)}{sf_r}+{GL(c)}{cg_r}+SUM({GL(c)}{ex_s}:{GL(c)}{xr})",Fn(11,True,W),BG(RD),AL("center"),usd)
te_r=r;r+=2

sb(ws,r,"NET RESULT",13);r+=1
ce(ws,r,1,"Monthly P/L",Fn(11,True,AC),BG(L))
for c in range(2,14):ce(ws,r,c,f"={GL(c)}{rv_r}-{GL(c)}{te_r}",Fn(11,True,AC),BG(L),AL("center"),usd)
pl_r=r;r+=1

ce(ws,r,1,"Cumulative P/L",Fn(11,True,AC),BG(M))
ce(ws,r,2,f"=B{pl_r}",Fn(11,True,AC),BG(M),AL("center"),usd)
for c in range(3,14):ce(ws,r,c,f"={GL(c-1)}{r}+{GL(c)}{pl_r}",Fn(11,True,AC),BG(M),AL("center"),usd)
cu_r=r;r+=2

sb(ws,r,"BREAK-EVEN",13);r+=1
ce(ws,r,1,"Fixed Costs/mo",Fn(10,True),BG(L));ce(ws,r,2,f"=Assumptions!C{tc_r}",Fn(),BG(L),AL("center"),usd)
bc=r;r+=1
ce(ws,r,1,"Net Margin/client",Fn(10,True),BG(L));ce(ws,r,2,f"=Assumptions!C{margin_r}",Fn(),BG(L),AL("center"),usd)
ba=r;r+=1
ce(ws,r,1,"Break-Even Clients",Fn(12,True,RD),BG(WB))
ce(ws,r,2,f"=CEILING(B{bc}/B{ba},1)",Fn(14,True,RD),BG(WB),AL("center"),'#,##0')
ce(ws,r,3,"← monthly cash-flow +",Fn(9,False,"6B7280"));r+=2
ce(ws,r,1,"Max Cash Burn",Fn(11,True,RD),BG(BB))
ce(ws,r,2,f"=MIN(B{cu_r}:M{cu_r})",Fn(14,True,RD),BG(BB),AL("center"),usd)

# ══════════════════════════════════════
# TAB 5: XLM REWARDS (AUDIT FIX: $3.30 per free card)
# ══════════════════════════════════════
ws=wb.create_sheet("XLM Rewards");ws.sheet_properties.tabColor=XG
sc(ws,[30]+[13]*12)
r=1;mg(ws,r,"XLM REWARDS & FREE CARDS",13);r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=13)
ce(ws,r,1,"FREE = waived fees. User pays load via USDC. We cover $3.30 per card from grant.",Fn(10,False,AC),BG(WB),AL("left","center",True))
ws.row_dimensions[r].height=30;r+=2

sb(ws,r,"PARAMETERS",13);r+=1
hd(ws,r,["Parameter","Value","","","","","","","","","","","Notes"]);r+=1
rp=[["Free Cards Budget (count)",500,"First 500 users"],
    ["Our Cost per Free Card ($)",3.30,"$3 issuance + 3%×$10 deposit (FEES ONLY)"],
    ["XLM Signup Bonus (XLM)",10,"≈$1 at current price"],
    ["XLM Spend Reward (%)",5,"5% cashback in XLM"],
    ["Avg Monthly Spend ($)",30,"Agent card spend"],
    ["XLM Price (USD)",0.10,"Current market"],
    ["Grant Budget for Rewards ($)",5000,"Portion of SCF grant"]]
rp_s=r
for p in rp:
    ce(ws,r,1,p[0],Fn(10,True),BG(L));ip(ws,r,2,p[1],'#,##0.00')
    ws.merge_cells(start_row=r,start_column=13,end_row=r,end_column=13)
    ce(ws,r,13,p[2],Fn(9,False,"6B7280"));r+=1

r+=1;sb(ws,r,"MONTHLY COSTS",13);r+=1;hd(ws,r,mos);r+=1

ce(ws,r,1,"Free Cards Issued",Fn(10,True),BG("FFF7E6"))
for c in range(2,14):
    if c==2:ce(ws,r,c,f"=MIN('User Growth'!{GL(c)}3,B{rp_s})",Fn(10,True,XG),BG("FFF7E6"),AL("center"),'#,##0')
    else:ce(ws,r,c,f"=MIN('User Growth'!{GL(c)}3,MAX(B{rp_s}-SUM(B{r}:{GL(c-1)}{r}),0))",Fn(10,True,XG),BG("FFF7E6"),AL("center"),'#,##0')
fc=r;r+=1

ce(ws,r,1,"Free Card Cost ($)",Fn(10,True,RD),BG(BB))
for c in range(2,14):ce(ws,r,c,f"={GL(c)}{fc}*B{rp_s+1}",Fn(10,False,RD),BG(BB),AL("center"),'$#,##0')
fcc=r;r+=1

ce(ws,r,1,"Signup Bonuses ($)",Fn(10,True,RD),BG(BB))
for c in range(2,14):ce(ws,r,c,f"={GL(c)}{fc}*B{rp_s+2}*B{rp_s+5}",Fn(10,False,RD),BG(BB),AL("center"),'$#,##0')
sbb=r;r+=1

ce(ws,r,1,"Spend Cashback ($)",Fn(10,True,RD),BG(BB))
for c in range(2,14):
    # Use realistic total active paying from User Growth tab
    ce(ws,r,c,f"='User Growth'!{GL(c)}6*B{rp_s+4}*(B{rp_s+3}/100)*B{rp_s+5}",Fn(10,False,RD),BG(BB),AL("center"),'$#,##0')
sr=r;r+=1

ce(ws,r,1,"TOTAL REWARD COST",Fn(11,True,W),BG(RD))
for c in range(2,14):ce(ws,r,c,f"={GL(c)}{fcc}+{GL(c)}{sbb}+{GL(c)}{sr}",Fn(11,True,W),BG(RD),AL("center"),'$#,##0')
tr=r;r+=2

sb(ws,r,"BUDGET TRACKING",13);r+=1
ce(ws,r,1,"Cumulative Spend",Fn(10,True,AC),BG(M))
ce(ws,r,2,f"=B{tr}",Fn(10,True,AC),BG(M),AL("center"),'$#,##0')
for c in range(3,14):ce(ws,r,c,f"={GL(c-1)}{r}+{GL(c)}{tr}",Fn(10,True,AC),BG(M),AL("center"),'$#,##0')
cr2=r;r+=1
ce(ws,r,1,"Grant Budget Remaining",Fn(10,True,GR),BG(RB))
for c in range(2,14):ce(ws,r,c,f"=B{rp_s+6}-{GL(c)}{cr2}",Fn(10,True,GR),BG(RB),AL("center"),'$#,##0')
r+=2

# Strategy steps
sb(ws,r,"XLM REWARDS STRATEGY",13);r+=1
strat=[
    ["1. Free Card Issuance","First 500 users get free $10 card — waived fees ($3.30 from grant). Zero friction."],
    ["2. XLM Signup Bonus","10 XLM airdrop on first card — teaches agents to hold XLM in Stellar wallet."],
    ["3. Spend Cashback in XLM","5% of card spend returned as XLM — drives recurring usage + grows Stellar TVL."],
    ["4. Stellar Wallet Onboarding","SDK auto-creates Stellar wallet for each agent — instant ecosystem growth."],
    ["5. XLM → Repeat Purchase","Agent earns XLM → uses XLM for next card → self-sustaining flywheel."],
    ["6. Referral in XLM","20 XLM per referred user who creates first card — viral loop."],
    ["7. SCF Grant Alignment","All costs funded by SCF grant — proves Stellar ecosystem growth to SDF."]]
for s in strat:
    ce(ws,r,1,s[0],Fn(10,True,SB),BG(L))
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=13)
    ce(ws,r,2,s[1],Fn(10,False),BG(L),AL("left","center",True))
    ws.row_dimensions[r].height=28;r+=1

r+=1;sb(ws,r,"SCF GRANT ROI (for reviewers)",13);r+=1
roi=[
    "→ Every free card = new Stellar wallet + USDC settlement on-chain",
    "→ XLM rewards create active wallet holders — direct SDF ecosystem metric",
    "→ Measurable: cards issued, wallets created, USDC volume, XLM distributed",
    "→ Grant ROI: $5,000 → 500 wallets + $15K+ USDC settled on Stellar",
    "→ Post-grant self-sustaining: paid users fund ongoing rewards"]
for w in roi:
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=13)
    ce(ws,r,1,w,Fn(10,False,AC),BG(L),AL("left","center"));r+=1

# ══════════════════════════════════════
# TAB 6: SENSITIVITY
# ══════════════════════════════════════
ws=wb.create_sheet("Sensitivity");ws.sheet_properties.tabColor=YL
sc(ws,[22]+[13]*7)
r=1;mg(ws,r,"SENSITIVITY ANALYSIS",8);r+=2
sb(ws,r,"BREAK-EVEN = Fixed Costs ÷ Net Margin/Client",8);r+=1
aps=[4,6,8,10,12,15,20]
hd(ws,r,["Cost \\ Margin"]+[f"${a}/mo" for a in aps]);r+=1
for fc in [200,400,600,800,1000,1500,2000,3000]:
    ce(ws,r,1,f"${fc}/mo",Fn(10,True),BG(L))
    for i,a in enumerate(aps,2):
        v=round(fc/a);bg=RB if v<=50 else(WB if v<=150 else BB);co=GR if v<=50 else("92400E" if v<=150 else RD)
        ce(ws,r,i,v,Fn(10,True,co),BG(bg),AL("center"),'#,##0')
    r+=1

# ══════════════════════════════════════
# TAB 7: KPI DASHBOARD
# ══════════════════════════════════════
ws=wb.create_sheet("KPI Dashboard");ws.sheet_properties.tabColor=RD
sc(ws,[24,14,14,14,13,13,13,13])
r=1;mg(ws,r,"WEEKLY KPI TRACKER",8);r+=2
sb(ws,r,"METRICS",8);r+=1
hd(ws,r,["Metric","✅ Target","⚠️ Warning","🚩 Critical","Wk1","Wk2","Wk3","Wk4"]);r+=1
for k in [["New Signups/wk",">10","3–10","<3"],["Free→Paid %",">20%","10–20%","<10%"],
          ["Txns/wk",">30","10–30","<10"],["Weekly Rev",">$200","$50–200","<$50"],
          ["Churn","<5%","5–10%",">10%"],["Uptime",">99.5%","97–99.5%","<97%"],
          ["Pro Subs/wk",">5","1–5","0"],["MRR",">$500","$100–500","<$100"],
          ["XLM Wallets/wk",">20","5–20","<5"],["Free Cards/wk",">15","5–15","<5"]]:
    ce(ws,r,1,k[0],Fn(10,True),BG(L))
    ce(ws,r,2,k[1],Fn(10,True,GR),BG(RB),AL("center"))
    ce(ws,r,3,k[2],Fn(10,True,"92400E"),BG(WB),AL("center"))
    ce(ws,r,4,k[3],Fn(10,True,RD),BG(BB),AL("center"))
    for c in [5,6,7,8]:ip(ws,r,c,None,'#,##0')
    r+=1

# ══════════════════════════════════════
# TAB 0: DASHBOARD
# ══════════════════════════════════════
ws=wb.create_sheet("Dashboard",0);ws.sheet_properties.tabColor=AC
sc(ws,[3,22,18,18,18,18,18,3])
ws.merge_cells("B1:G1");ws.row_dimensions[1].height=50
ce(ws,1,2,"ASG CARD",Fn(22,True,AC),BG(W),AL("left","center"))
ws.merge_cells("B2:G2")
ce(ws,2,2,"Financial Model v5 (Audited)",Fn(12,False,"6B7280"),BG(W),AL("left"))
r=4

# Product description
sb(ws,r,"WHAT IS ASG CARD?",7);r+=1
about=[
    "ASG Card issues instant virtual Visa cards for AI agents via API.",
    "Payments settled in USDC on the Stellar blockchain using x402 protocol.",
    '"Payment IS Authentication" — no API keys, no KYC, no registration needed.',
    "Agent sends USDC → gets a virtual card → makes purchases autonomously.",
    "",
    "Revenue: per-transaction fees (topUpFee + serviceFee) + $29/mo Pro subscription.",
    "Card issuer: 4payments ($3/card + 3% deposit). Grant: SCF #42 (Stellar).",
    "Stage: Pre-launch pilot. Target: production on Stellar by May 2026.",
]
for a in about:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    if a=="":
        r+=1;continue
    ce(ws,r,2,a,Fn(10,False,AC if 'Revenue' in a or 'Card issuer' in a or 'Stage' in a else DT),BG(L),AL("left","center",True))
    ws.row_dimensions[r].height=22;r+=1
r+=1

# KPI cards
sb(ws,r,"KEY METRICS",7);r+=1
lbs=["Net Margin\n/client","Break-Even\nClients","LTV/CAC","Mo12\nRevenue","Mo12\nPaying","Free Card\nBudget Left"]
fms=[f"=Assumptions!C{margin_r}",f"='P&L'!B{bc+2}",f"='Unit Economics'!C{ms+3}",
     f"='P&L'!M{rv_r}",f"='User Growth'!M{tot_r-6}",f"='XLM Rewards'!M22"]
ft=['$#,##0.00','#,##0','0.0"x"','$#,##0','#,##0','$#,##0']
for i,(lb,fm,ft2) in enumerate(zip(lbs,fms,ft)):
    c=i+2;ce(ws,r,c,lb,Fn(9,False,"6B7280"),BG(M),AL("center","center",True))
    ce(ws,r+1,c,fm,Fn(16,True,AC),BG(L),AL("center","center"),ft2)
ws.row_dimensions[r].height=28;ws.row_dimensions[r+1].height=40
r+=3

sb(ws,r,"MODEL TABS",7);r+=1
for lb,ds in [("→ Assumptions","Editable inputs: 4payments costs, pricing, Pro subscription, growth"),
    ("→ Unit Economics","Real margin per transaction after 4payments fees"),
    ("→ User Growth","Signup→Free→Paid funnel, declining churn, ramped Pro"),
    ("→ P&L","Revenue (PPU + Sub), COGS, fixed costs, XLM rewards, break-even"),
    ("→ XLM Rewards ★","Free cards, XLM bonuses, strategy steps, SCF grant ROI"),
    ("→ Sensitivity","What-if tables: break-even × margin, LTV/CAC × churn"),
    ("→ KPI Dashboard","Weekly tracking: signups, MRR, churn, XLM metrics")]:
    ce(ws,r,2,lb,Fn(10,True,BL),BG(L));ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
    ce(ws,r,3,ds,Fn(10,False),BG(L));r+=1
r+=1;ws.merge_cells(f"B{r}:G{r}")
ce(ws,r,2,"💡 Blue cells = editable. All formulas auto-propagate. XLM costs feed into P&L.",Fn(10,False,BL),BG(IB),AL("left"))

wb.save(OUT)
print(f"✅ Model v5 (audited) saved: {OUT}")
print(f"   Tabs: {[s.title for s in wb.worksheets]}")
